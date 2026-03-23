"""Tests for XLSX exporter."""
import io
from openpyxl import load_workbook
from app.export.xlsx_exporter import export_spreadsheet


def _open(data: bytes):
    return load_workbook(io.BytesIO(data))


def test_export_cell_values():
    meta = {"sheets": [{"name": "Sheet1"}], "cells": {"Sheet1": {"A1": {"value": "Hello"}, "B1": {"value": 42}}}}
    wb = _open(export_spreadsheet(meta, "Test"))
    ws = wb.active
    assert ws["A1"].value == "Hello"
    assert ws["B1"].value == 42

def test_export_formula():
    meta = {"sheets": [{"name": "Sheet1"}], "cells": {"Sheet1": {"A1": {"value": 10, "formula": "SUM(B1:B5)"}}}}
    wb = _open(export_spreadsheet(meta, "Test"))
    assert wb.active["A1"].value == "=SUM(B1:B5)"

def test_export_bold():
    meta = {"sheets": [{"name": "Sheet1"}], "cells": {"Sheet1": {"A1": {"value": "Bold", "format": {"bold": True}}}}}
    wb = _open(export_spreadsheet(meta, "Test"))
    assert wb.active["A1"].font.bold

def test_export_multiple_sheets():
    meta = {"sheets": [{"name": "Data"}, {"name": "Summary"}], "cells": {"Data": {"A1": {"value": 1}}, "Summary": {"A1": {"value": 2}}}}
    wb = _open(export_spreadsheet(meta, "Test"))
    assert wb.sheetnames == ["Data", "Summary"]
    assert wb["Data"]["A1"].value == 1
    assert wb["Summary"]["A1"].value == 2

def test_export_valid_xlsx():
    meta = {"sheets": [{"name": "Sheet1"}], "cells": {"Sheet1": {"A1": {"value": "test"}}}}
    data = export_spreadsheet(meta, "Test")
    assert len(data) > 100
    import zipfile
    zf = zipfile.ZipFile(io.BytesIO(data))
    assert "xl/workbook.xml" in zf.namelist()

def test_export_empty():
    meta = {"sheets": [{"name": "Sheet1"}], "cells": {}}
    data = export_spreadsheet(meta, "Test")
    wb = _open(data)
    assert wb.active.title == "Sheet1"

def test_export_column_widths():
    meta = {"sheets": [{"name": "Sheet1"}], "cells": {}, "columnWidths": {"Sheet1": {"A": 20}}}
    wb = _open(export_spreadsheet(meta, "Test"))
    assert wb.active.column_dimensions["A"].width == 20
