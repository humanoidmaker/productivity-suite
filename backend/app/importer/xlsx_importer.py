from __future__ import annotations
"""Import XLSX to internal spreadsheet format using openpyxl."""

import io

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def import_xlsx(data: bytes) -> dict:
    """Convert XLSX bytes to internal spreadsheet meta dict."""
    wb = load_workbook(io.BytesIO(data), data_only=False)

    sheets_info = []
    cells_data = {}
    col_widths = {}
    merged_cells_map = {}
    freeze_panes = {}

    for idx, ws in enumerate(wb.worksheets):
        sheet_name = ws.title
        sheets_info.append({"name": sheet_name, "index": idx, "visible": ws.sheet_state == "visible", "color": None})

        sheet_cells = {}
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    ref = f"{get_column_letter(cell.column)}{cell.row}"
                    cell_info: dict = {"value": cell.value}
                    if cell.data_type == "f":
                        cell_info["formula"] = str(cell.value).lstrip("=") if str(cell.value).startswith("=") else str(cell.value)
                    fmt = {}
                    if cell.font.bold:
                        fmt["bold"] = True
                    if cell.font.italic:
                        fmt["italic"] = True
                    if cell.font.name:
                        fmt["fontFamily"] = cell.font.name
                    if cell.font.size:
                        fmt["fontSize"] = cell.font.size
                    if cell.fill.fgColor and cell.fill.fgColor.rgb and cell.fill.fgColor.rgb != "00000000":
                        fmt["backgroundColor"] = f"#{cell.fill.fgColor.rgb[2:]}" if len(cell.fill.fgColor.rgb) > 6 else f"#{cell.fill.fgColor.rgb}"
                    if cell.alignment.horizontal:
                        fmt["textAlign"] = cell.alignment.horizontal
                    if fmt:
                        cell_info["format"] = fmt
                    sheet_cells[ref] = cell_info

        cells_data[sheet_name] = sheet_cells

        # Column widths
        widths = {}
        for col_dim in ws.column_dimensions.values():
            if col_dim.width and col_dim.width != 8.0:  # 8.0 is default
                letter = get_column_letter(col_dim.index) if hasattr(col_dim, "index") else col_dim.letter
                if letter:
                    widths[letter] = col_dim.width
        if widths:
            col_widths[sheet_name] = widths

        # Merged cells
        merges = [str(m) for m in ws.merged_cells.ranges]
        if merges:
            merged_cells_map[sheet_name] = merges

        # Freeze panes
        if ws.freeze_panes:
            freeze_panes[sheet_name] = str(ws.freeze_panes)

    return {
        "sheets": sheets_info,
        "cells": cells_data,
        "columnWidths": col_widths,
        "mergedCells": merged_cells_map,
        "freezePanes": freeze_panes,
    }
