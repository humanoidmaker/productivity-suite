from __future__ import annotations
"""Export spreadsheet data to XLSX using openpyxl."""

import io
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def export_spreadsheet(meta: dict, title: str) -> bytes:
    """Convert internal spreadsheet data to XLSX bytes."""
    wb = Workbook()

    sheets_info = meta.get("sheets", [{"name": "Sheet1", "index": 0}])
    cells_data = meta.get("cells", {})

    for i, sheet_info in enumerate(sheets_info):
        if i == 0:
            ws = wb.active
            ws.title = sheet_info.get("name", "Sheet1")
        else:
            ws = wb.create_sheet(title=sheet_info.get("name", f"Sheet{i + 1}"))

        sheet_name = sheet_info.get("name", "Sheet1")
        sheet_cells = cells_data.get(sheet_name, {})

        for cell_ref, cell_data in sheet_cells.items():
            cell = ws[cell_ref]
            if isinstance(cell_data, dict):
                cell.value = cell_data.get("value")
                if cell_data.get("formula"):
                    cell.value = f"={cell_data['formula']}"
                fmt = cell_data.get("format", {})
                if fmt.get("bold") or fmt.get("italic"):
                    cell.font = Font(bold=fmt.get("bold", False), italic=fmt.get("italic", False), name=fmt.get("fontFamily", "Calibri"), size=fmt.get("fontSize", 11))
                if fmt.get("backgroundColor"):
                    cell.fill = PatternFill(start_color=fmt["backgroundColor"].lstrip("#"), fill_type="solid")
                if fmt.get("textAlign"):
                    cell.alignment = Alignment(horizontal=fmt["textAlign"])
            else:
                cell.value = cell_data

        # Column widths
        col_widths = meta.get("columnWidths", {}).get(sheet_name, {})
        for col_letter, width in col_widths.items():
            ws.column_dimensions[col_letter].width = width

        # Merged cells
        merged = meta.get("mergedCells", {}).get(sheet_name, [])
        for merge_range in merged:
            ws.merge_cells(merge_range)

        # Freeze panes
        freeze = meta.get("freezePanes", {}).get(sheet_name)
        if freeze:
            ws.freeze_panes = freeze

    wb.properties.title = title

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
